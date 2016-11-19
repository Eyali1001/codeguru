from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

scoresdir ="C:\\corewars2\\Corewars\\scores.csv" #CHANGE TO YOUR CORRECT DIRECTORY

f = open(scoresdir,'r').read()
f = ",".join(f.splitlines()).split(",")
a = [(f[i],f[i+1]) for i in range(0,len(f)-1,2)]
a = list(set(a))
a = sorted(a,key=lambda x: float(x[1]))
a.reverse()
print a
wb = Workbook()
dest_filename = 'scores.xlsx'
ws1 = wb.active
ws1.title = "scores"

for row in range(1,len(a)+1):
	for col in [1,2]:
		 _ = ws1.cell(column=col, row=row, value=a[row-1][col-1])
		 
wb.save(filename = dest_filename)